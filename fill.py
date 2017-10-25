import openpyxl
import selenium
from selenium import webdriver
import time
driver=webdriver.Firefox()
driver.get("http://192.168.4.19:8886/Nichehire/")
book=openpyxl.load_workbook('/home/keertmak/login.xlsx')
sheet=book.active
cell=sheet['A1':'B3']
x=sheet.cell(row=1,column=1).value
y=sheet.cell(row=1,column=2).value
dr=driver.find_element_by_id("email_id").send_keys(sheet.cell(row=1,column=1).value)
dr=driver.find_element_by_name("pwd").send_keys(sheet.cell(row=1,column=2).value)
dr=driver.find_element_by_id("submitBtn").click()
